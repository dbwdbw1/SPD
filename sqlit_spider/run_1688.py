import datetime
import re
import sys
import time

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By

import consts
from sqlit_spider.sqlite3_model import SessionContext, GoodsInfo


def driver_init():
    # 加载配置数据
    profile_dir = r'--user-data-dir=C:\Users\hjl\AppData\Local\Google\Chrome\User Data'
    c_option = webdriver.ChromeOptions()
    c_option.add_experimental_option('useAutomationExtension', False)
    c_option.add_experimental_option('excludeSwitches', ['enable-automation'])

    prefs = {
        'profile.default_content_setting_values':
            {
                'notifications': 2
            },
        'profile.password_manager_enabled': False,
        'credentials_enable_service': False
    }
    c_option.add_experimental_option('prefs', prefs)

    # 开发者模式防止被识别出
    # 网址：https://blog.csdn.net/dslkfajoaijfdoj/article/details/109146051
    c_option.add_experimental_option('excludeSwitches', ['enable-automation'])
    c_option.add_argument("--disable-blink-features=AutomationControlled")
    # c_option.add_experimental_option('w3c', False)

    c_option.add_argument(profile_dir)
    driver = webdriver.Chrome(chrome_options=c_option)
    # 执行cdp命令
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
                    Object.defineProperty(navigator, 'webdriver', {
                        get: () => undefined
                    })
                  """
    })

    return driver


# 打开链接获取商品的详细信息
def get_goods_detail(driver, link, original_window):
    while True:
        try:
            # 打开新窗口
            driver.execute_script("window.open('" + link + "')")
            # 等待浏览器加载完毕
            driver.implicitly_wait(7)
            # 获取所有窗口句柄
            all_handles = driver.window_handles
            goods_detail_handle = ''
            # 切换到新窗口
            for handle in all_handles:
                if handle != original_window:
                    driver.switch_to.window(handle)
                    goods_detail_handle = handle
                    break
            time.sleep(2)
            aliwangwang_button = driver.find_element(By.XPATH, "//a[@target='_blank' and @class='tool-item']")
            # 先点击展开按钮
            try:
                driver.find_element(By.CLASS_NAME, 'offer-attr-switch').click()
            except:
                pass
            time.sleep(1)

            try:
                # 获取商品信息
                goods_div = driver.find_elements(By.XPATH,
                                                 '//div[@class="offer-attr-list"]/div[@class="offer-attr-item"]')
            except Exception as e:
                print('商品信息解决出错，link:', link, 'err: ', e)
                goods_div = []
            try:
                # 获取地址信息
                address_div = driver.find_element(
                    By.XPATH,
                    '//div[@id="shop-container-footer"]//div[@style="text-align: center;"]/p'
                )
                factory_name = address_div.find_elements(By.TAG_NAME, 'span')[0].text
                address = address_div.find_elements(By.TAG_NAME, 'span')[1].text.replace('地址：', '')
            except Exception as e:
                print('地址信息解决出错，link:', link, 'err: ', e)
                factory_name = '工厂名获取出错'
                address = '工厂名获取出错'
            # 获取厂家名称
            print(factory_name, address)

            print('-----------------')
            # 定义商品详情表的数据字典
            data = []
            for span in goods_div:
                name = span.find_element(By.CLASS_NAME, 'offer-attr-item-name').text.strip()
                value = span.find_element(By.CLASS_NAME, 'offer-attr-item-value').text.strip()
                item = {
                    'name': name,
                    'value': value
                }
                data.append(item)
            aliwangwang_button.click()
            try:
                driver.find_element(By.XPATH, '//button[text()="进入网页版"]').click()
                current_handle_index = driver.window_handles.index(driver.current_window_handle)
                driver.switch_to.window(driver.window_handles[current_handle_index + 1])
                driver.find_element(By.XPATH, '//span[@class="next-btn-helper" and text()="使用网页版"]').click()
            except Exception:
                print('没有弹窗选择旺旺版本，默认进入网页版')

            # 点完弹窗，切换回聊天窗口
            for handle in driver.window_handles:
                driver.switch_to.window(handle)
                if driver.title == '1688旺旺聊天':
                    break
            time.sleep(2)
            print('商品信息获取完毕')
            # 关闭聊天窗口
            driver.close()

            # 关闭商品详情窗口
            driver.switch_to.window(goods_detail_handle)
            driver.close()
            # 切换回原来的窗口
            driver.switch_to.window(original_window)
            # 返回商品详情表的数据字典
            return (str(data), address, factory_name)

        except ValueError as e:
            print("捕获到异常:", e)
            # 手动处理滑动验证
            input('请手动滑动验证码，输入任意字符继续：')
            continue


def get_goods_info(driver, original_window, origin_img_name):
    # 获取商品信息的列表
    goods_list = driver.find_element(By.ID, 'sm-offer-list').find_elements(By.CLASS_NAME, 'space-offer-card-box')
    truncated_goods_list = []
    if goods_list:
        truncated_goods_list = goods_list[:consts.GET_GOODS_INFO_COUNT]
    print('-----------------')
    print('开始爬取数据，数据长度为：', len(truncated_goods_list))
    # 创建一个工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.append(['商品链接', '商品信息', '截图'])
    ws.column_dimensions['A'].width = consts.GOODS_URL_CELL_WIDTH
    ws.column_dimensions['B'].width = consts.GOODS_TITLE_CELL_WIDTH
    goods_index = 0
    for li in truncated_goods_list:
        row = goods_index + 2  # Excel的行从1开始计数，所以除了表头，第一条数据应该是0 + 2
        link_div = li.find_element(By.CLASS_NAME, 'mojar-element-title').find_element(By.TAG_NAME, 'a')
        link = link_div.get_attribute('href')
        title = li.find_element(By.CLASS_NAME, 'mojar-element-title').find_element(By.CLASS_NAME, 'title').text
        price_div = li.find_element(By.CLASS_NAME, 'mojar-element-price')
        price = price_div.find_element(By.CLASS_NAME, 'showPricec').text
        sale_sum = price_div.find_element(By.CLASS_NAME, 'sale').text
        hyperlink = Hyperlink(ref=link)
        ws.cell(row=row, column=1, value=f"{link}").hyperlink = hyperlink
        ws.cell(row=row, column=1).font = Font(color="0000FF", underline="single")  # 设置字体颜色和下划线
        ws.cell(row=row, column=2, value=title)
        # 获取商品图片链接匹配对象
        match_obj = re.search(r'url\("([^"]+)"\);', li.find_element(By.CLASS_NAME, 'img').get_attribute("style"))
        # 下载并保存图片
        if match_obj:
            img_url = match_obj.group(1)
            # 下载图片并保存
            img_response = requests.get(img_url, stream=True)
            if img_response.status_code == 200:
                img_data = img_response.content
                # 获取图片文件名
                img_filename = f"./resource/goods_picture/{goods_index}_image.jpg"
                with open(img_filename, 'wb') as img_file:
                    img_file.write(img_data)
                add_pic_to_excel(goods_index, img_filename, row, ws)

        if sale_sum == '':
            sale_sum = '0'
        print('-----------------')
        print(title, price, sale_sum)
        # 插入数据
        with SessionContext() as session:
            # 获取商品详细信息
            detail_data, address, factory_name = get_goods_detail(driver, link, original_window)
            # 插入数据
            goods = GoodsInfo(title=title, price=price, sale_sum=sale_sum, link=link, detail=detail_data,
                              address=address, factory_name=factory_name)
            session.add(goods)
            session.commit()
        goods_index += 1
    current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    wb.save(f"./resource/goods_info_{origin_img_name}_{current_time}.xlsx")


# 将图片插入到工作表
def add_pic_to_excel(goods_index, img_filename, row, ws):
    img = ExcelImage(img_filename)
    new_size = (consts.PIC_WIDTH, consts.PIC_HEIGHT)
    img.height, img.width = new_size
    ws.row_dimensions[row].height = consts.PIC_CELL_HEIGHT
    ws.column_dimensions['C'].width = consts.PIC_CELL_WIDTH
    img.anchor = f"C{goods_index + 2}"  # 设置图片位置，这里假设图片插入到每行数据的D列
    ws.add_image(img)


def check_login_redirect(web_driver):
    max_check_times = 10
    for i in range(max_check_times):
        if web_driver.current_url.startswith(consts.DEFAULT_1688_URL_PREFIX):
            return True
        else:
            time.sleep(2)
    return False


# 启动爬虫程序
def start_crawler(image2url):
    driver = driver_init()
    driver.implicitly_wait(3)

    for imageName, imageUrl in image2url.items():
        print("开始搜索图片：" + imageName)
        crawl_by_image_url(driver, imageName, imageUrl)
    # 关闭浏览器
    driver.quit()


# 根据图片url爬取数据
def crawl_by_image_url(driver, image_name, image_url):
    # 打开链接获取商品信息
    driver.get(consts.DEFAULT_1688_URL_PREFIX)
    if driver.current_url.startswith(consts.LOGIN_URL_PREFIX):
        time.sleep(3)
        if not check_login_redirect(driver):
            print('未检测到登录行为。请关闭浏览器并重启程序后进行登录以继续使用。')
            driver.quit()
            sys.exit()
    # 提供的 cookie 字符串
    cookie_string = 'mtop_partitioned_detect=1; _m_h5_tk=68e919303ae296e3590c8c90cc7ee29c_1707025982338; _m_h5_tk_enc=45c835786dda208e6d5def9ea18e5f24; lid=dbwdbw1; ali_ab=180.111.221.27.1707016441129.5; cna=qpnxHcLZDykCATFBUP4VLqTU; taklid=6a0748690e0048a996484d47c1ed1f53; mwb=ng; xlly_s=1; cookie2=1f0ac2b0f8f1b5955a0c41d9d918121d; _tb_token_=e3ef33abe5138; enc=lhjdhx3Ve3AgvYkwk%2BgOxZ84vZulpGk4K9%2Bf1LWKtrgQU1JDTXmJpkndHBV43C7vpZjk03vCVEeIyHCCCyHb0w%3D%3D; tracknick=dbwdbw1; cna=m9dLGyl9DE4CAbRmjwLdtIsy; inquiry_preference=webIM; cookie1=AVdHXHZcvIGj9hg1z5Z3ZebSVUEfwSNLF%2BDmf8WRKcc%3D; cookie17=UUplaXFNqNx8Lg%3D%3D; sgcookie=E1004wyhm7qXKHOX8nqZgmhlWoP%2B7MQ5sPoa%2B59BbWa6QuMu3aWXVXZ656NKtjbIfNkhaSi6I%2B69Ty1xIoX23EuDDtljWHvu5d3kP0fuLzSsSxbde6Fs6Vk3bzYMvCUaSUst; t=9c762f68f95bccbc5f663bdfa0e59289; sg=187; csg=29b2b728; unb=2248476908; uc4=nk4=0%40BQOkCq2f08%2BcO6u3hX8s%2Fnda&id4=0%40U2gvJul3CLHC35j%2FHiGTddjgo%2FsN; __cn_logon__=true; __cn_logon_id__=dbwdbw1; ali_apache_track=c_mid=b2b-2248476908|c_lid=dbwdbw1|c_ms=1; ali_apache_tracktmp=c_w_signed=Y; _nk_=dbwdbw1; last_mid=b2b-2248476908; _csrf_token=1707017096942; is_identity=buyer; __mwb_logon_id__=dbwdbw1; tfstk=e-Kpb_mQZfcHbXPVSD3M4a1ArAHGwHpFBBJbq_f3PCd9F1XQPuZWW_d666whx9m7VQAGOpvoaue56BZl-3pz2T9Wav5oLVveLgSSijf-mpJFIPp0UmDlawORVjcmjv3aMiSW-fGGbiNsgSlMn3TXfxOFqBm102xGpwC8q1ERtercRsNf_uZ599QLWp1TVut14HtDD20UisBuRAHTzz7CQ7DWVm3EpkPfBsDCRzzPrOWOiAHTzz7CQOCmduazzaXN.; isg=BFtbYbkDul1au8d_QmnUGssj6r_FMG8yttLOI02TV9pxLHoO1QJ4gDrtxoyiDMcq'

    # 将字符串分割成键值对
    cookies = [cookie.split("=") for cookie in cookie_string.split("; ")]

    for cookie in cookies:
        driver.add_cookie({'name': cookie[0], 'value': cookie[1], 'domain': '1688.com'})

    driver.get(image_url)
    # 睡眠3秒，等待浏览器加载完毕
    time.sleep(3)

    # TODO 价格排序后，前面的很可能不是符合图片的，最好还是默认“综合排序”
    # # 按价格排序
    # price_sort_button = driver.find_element(By.XPATH, "//span[@class='sm-widget-txt' and text()='价格']")
    # price_sort_button.click()

    # 等待浏览器加载完毕
    driver.implicitly_wait(7)
    # 获取当前窗口句柄
    original_window = driver.current_window_handle
    time.sleep(1)
    # 慢慢向下滚动
    for i in range(1, 3):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)
    # 再慢慢向上滚动
    for i in range(1, 3):
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(1)
    # 滑动鼠标到页面中部
    ActionChains(driver).move_by_offset(0, 300).perform()
    time.sleep(1)

    # 获取商品信息
    get_goods_info(driver, original_window, image_name[9:])
    print('-----------------')
    print(image_name + '爬取完毕！')
